"""
Reserve Adequacy Quick Check — Main Analysis Engine

Runs standard actuarial reserving methods on loss development triangles
and produces a formatted Excel report with diagnostics.
"""
import pandas as pd
import numpy as np
import argparse
import json
import sys
import os

# Add parent dir to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.parse_triangle import read_triangle, read_premium


# =============================================================================
# Core Actuarial Computations
# =============================================================================

def compute_ata_factors(triangle, method="volume_weighted"):
    """
    Compute age-to-age (link ratio) factors.
    Returns DataFrame of individual factors and selected factors.
    """
    n_rows, n_cols = triangle.shape
    dev_periods = triangle.columns.tolist()
    individual_factors = pd.DataFrame(index=range(n_rows), columns=dev_periods[:-1])

    for col_idx in range(n_cols - 1):
        d1 = dev_periods[col_idx]
        d2 = dev_periods[col_idx + 1]
        for row_idx in range(n_rows):
            val1 = triangle.iloc[row_idx, col_idx]
            val2 = triangle.iloc[row_idx, col_idx + 1]
            if pd.notna(val1) and pd.notna(val2) and val1 != 0:
                individual_factors.iloc[row_idx, col_idx] = val2 / val1

    individual_factors = individual_factors.apply(pd.to_numeric, errors='coerce')

    selected = {}
    for col_idx in range(n_cols - 1):
        d1 = dev_periods[col_idx]
        d2 = dev_periods[col_idx + 1]
        col_factors = individual_factors.iloc[:, col_idx].dropna()

        if len(col_factors) == 0:
            selected[d1] = {'volume_weighted': np.nan, 'simple_avg': np.nan,
                           'medial_avg': np.nan, 'min': np.nan, 'max': np.nan,
                           'std': np.nan, 'count': 0}
            continue

        # Volume-weighted
        numerator = 0.0
        denominator = 0.0
        for row_idx in range(n_rows):
            val1 = triangle.iloc[row_idx, col_idx]
            val2 = triangle.iloc[row_idx, col_idx + 1]
            if pd.notna(val1) and pd.notna(val2) and val1 != 0:
                numerator += val2
                denominator += val1
        vw = numerator / denominator if denominator != 0 else np.nan

        # Simple average
        sa = col_factors.mean()

        # Medial average (exclude high and low)
        if len(col_factors) >= 3:
            sorted_f = col_factors.sort_values()
            medial = sorted_f.iloc[1:-1].mean()
        else:
            medial = sa

        selected[d1] = {
            'volume_weighted': vw,
            'simple_avg': sa,
            'medial_avg': medial,
            'min': col_factors.min(),
            'max': col_factors.max(),
            'std': col_factors.std() if len(col_factors) > 1 else 0,
            'count': len(col_factors)
        }

    return individual_factors, selected


def compute_tail_factor(selected_factors, dev_periods):
    """Estimate tail factor using exponential decay extrapolation."""
    # Get the last few known factors
    factors = []
    periods = []
    for d in dev_periods[:-1]:
        f = selected_factors.get(d, {}).get('volume_weighted', np.nan)
        if pd.notna(f) and f > 1.0:
            factors.append(f - 1.0)
            periods.append(d)

    if len(factors) < 3:
        return 1.000, "insufficient_data"

    # Use last known factors for extrapolation
    last_n = min(5, len(factors))
    y = np.log(np.array(factors[-last_n:]))
    x = np.array(periods[-last_n:], dtype=float)

    # Check if factors are declining
    if not all(y[i] >= y[i+1] for i in range(len(y)-1)):
        # Not monotonically declining — use 1.000
        return 1.000, "non_declining"

    try:
        coeffs = np.polyfit(x, y, 1)
        if coeffs[0] >= 0:
            return 1.000, "non_declining_fit"

        # Extrapolate for additional periods
        tail = 1.0
        max_periods = dev_periods[-1]
        for extra in range(1, 100):
            projected = np.exp(coeffs[0] * (max_periods + extra) + coeffs[1])
            if projected < 0.0001:
                break
            tail *= (1.0 + projected)
        return round(tail, 4), "exponential_decay"
    except Exception:
        return 1.000, "fit_failed"


def compute_cdfs(selected_factors, dev_periods, tail_factor, method_key="volume_weighted"):
    """Compute cumulative development factors from age-to-age factors."""
    n = len(dev_periods)
    cdfs = {}
    cdfs[dev_periods[-1]] = tail_factor

    for i in range(n - 2, -1, -1):
        d = dev_periods[i]
        ata = selected_factors.get(d, {}).get(method_key, np.nan)
        if pd.notna(ata):
            cdfs[d] = ata * cdfs[dev_periods[i + 1]]
        else:
            cdfs[d] = cdfs[dev_periods[i + 1]]

    return cdfs


def chain_ladder_ultimates(triangle, acc_periods, cdfs, dev_periods):
    """Compute chain ladder ultimate losses and IBNR."""
    n_rows = len(acc_periods)
    results = []
    for i in range(n_rows):
        # Find latest non-NaN value
        latest_val = np.nan
        latest_dev = None
        for j in range(len(dev_periods) - 1, -1, -1):
            val = triangle.iloc[i, j]
            if pd.notna(val):
                latest_val = val
                latest_dev = dev_periods[j]
                break

        if pd.notna(latest_val) and latest_dev is not None:
            cdf = cdfs.get(latest_dev, 1.0)
            ultimate = latest_val * cdf
            ibnr = ultimate - latest_val
            pct_reported = 1.0 / cdf if cdf != 0 else np.nan
        else:
            ultimate = np.nan
            ibnr = np.nan
            pct_reported = np.nan
            latest_val = np.nan

        results.append({
            'accident_period': acc_periods[i],
            'latest_loss': latest_val,
            'latest_dev': latest_dev,
            'cdf': cdfs.get(latest_dev, np.nan) if latest_dev else np.nan,
            'ultimate': ultimate,
            'ibnr': ibnr,
            'pct_reported': pct_reported
        })

    return pd.DataFrame(results)


def bornhuetter_ferguson(triangle, acc_periods, dev_periods, cdfs, premiums, elr):
    """Compute BF ultimate losses."""
    results = []
    for i, ap in enumerate(acc_periods):
        latest_val = np.nan
        latest_dev = None
        for j in range(len(dev_periods) - 1, -1, -1):
            val = triangle.iloc[i, j]
            if pd.notna(val):
                latest_val = val
                latest_dev = dev_periods[j]
                break

        prem = premiums.get(str(ap), np.nan)
        cdf = cdfs.get(latest_dev, 1.0) if latest_dev else np.nan

        if pd.notna(latest_val) and pd.notna(prem) and pd.notna(cdf) and cdf != 0:
            pct_unreported = 1.0 - 1.0 / cdf
            bf_ibnr = elr * prem * pct_unreported
            ultimate = latest_val + bf_ibnr
        else:
            bf_ibnr = np.nan
            ultimate = np.nan

        results.append({
            'accident_period': ap,
            'latest_loss': latest_val,
            'premium': prem,
            'elr': elr,
            'pct_unreported': 1.0 - 1.0 / cdf if pd.notna(cdf) and cdf != 0 else np.nan,
            'bf_ibnr': bf_ibnr,
            'ultimate': ultimate
        })

    return pd.DataFrame(results)


def cape_cod(triangle, acc_periods, dev_periods, cdfs, premiums):
    """Compute Cape Cod (Stanard-Bühlmann) ultimate losses."""
    # Derive ELR from data
    numerator = 0.0
    denominator = 0.0

    for i, ap in enumerate(acc_periods):
        latest_val = np.nan
        latest_dev = None
        for j in range(len(dev_periods) - 1, -1, -1):
            val = triangle.iloc[i, j]
            if pd.notna(val):
                latest_val = val
                latest_dev = dev_periods[j]
                break

        prem = premiums.get(str(ap), np.nan)
        cdf = cdfs.get(latest_dev, 1.0) if latest_dev else np.nan

        if pd.notna(latest_val) and pd.notna(prem) and pd.notna(cdf) and cdf != 0:
            numerator += latest_val
            denominator += prem * (1.0 / cdf)

    elr_cc = numerator / denominator if denominator != 0 else np.nan

    # Now apply BF formula with derived ELR
    results = bornhuetter_ferguson(triangle, acc_periods, dev_periods, cdfs, premiums, elr_cc)
    results = results.rename(columns={'elr': 'cape_cod_elr', 'bf_ibnr': 'cc_ibnr'})
    results['cape_cod_elr'] = elr_cc

    return results, elr_cc


# =============================================================================
# Diagnostics
# =============================================================================

def calendar_year_test(triangle, acc_periods, dev_periods):
    """Test for consistency of development along calendar year diagonals."""
    n_rows, n_cols = triangle.shape
    diag_sums = {}

    for i in range(n_rows):
        for j in range(n_cols):
            val = triangle.iloc[i, j]
            if pd.notna(val):
                cal_year = i + j  # diagonal index
                if cal_year not in diag_sums:
                    diag_sums[cal_year] = 0.0
                diag_sums[cal_year] += val

    sorted_diags = sorted(diag_sums.items())
    results = []
    for idx in range(1, len(sorted_diags)):
        prev_key, prev_val = sorted_diags[idx - 1]
        curr_key, curr_val = sorted_diags[idx]
        if prev_val != 0:
            change = (curr_val - prev_val) / prev_val
        else:
            change = np.nan
        results.append({
            'diagonal': curr_key,
            'sum': curr_val,
            'prior_sum': prev_val,
            'change': change
        })

    return pd.DataFrame(results)


def outlier_factors(individual_factors, selected_factors, dev_periods):
    """Identify outlier age-to-age factors (>2 std dev from mean)."""
    outliers = []
    for col_idx in range(individual_factors.shape[1]):
        d = dev_periods[col_idx] if col_idx < len(dev_periods) - 1 else None
        if d is None:
            continue
        col = individual_factors.iloc[:, col_idx].dropna()
        if len(col) < 3:
            continue
        mean = col.mean()
        std = col.std()
        if std == 0:
            continue
        for row_idx in col.index:
            val = col[row_idx]
            if abs(val - mean) > 2 * std:
                outliers.append({
                    'dev_period': d,
                    'row_index': row_idx,
                    'factor': val,
                    'mean': mean,
                    'std_devs': (val - mean) / std
                })
    return pd.DataFrame(outliers) if outliers else pd.DataFrame()


def tail_sensitivity(cl_results, cdfs, tail_factor, dev_periods):
    """Compute ultimate sensitivity to tail factor changes."""
    scenarios = {
        'Selected': tail_factor,
        '-25%': 1.0 + (tail_factor - 1.0) * 0.75,
        '-10%': 1.0 + (tail_factor - 1.0) * 0.90,
        '+10%': 1.0 + (tail_factor - 1.0) * 1.10,
        '+25%': 1.0 + (tail_factor - 1.0) * 1.25,
    }

    results = {}
    for label, tf in scenarios.items():
        total_ult = 0.0
        total_latest = 0.0
        for _, row in cl_results.iterrows():
            if pd.notna(row['ultimate']) and pd.notna(row['latest_loss']):
                # Adjust ultimate for tail change
                if row['cdf'] != 0 and tail_factor != 0:
                    adjusted_cdf = row['cdf'] * tf / tail_factor
                    adj_ultimate = row['latest_loss'] * adjusted_cdf
                else:
                    adj_ultimate = row['ultimate']
                total_ult += adj_ultimate
                total_latest += row['latest_loss']
        results[label] = {
            'tail_factor': tf,
            'total_ultimate': total_ult,
            'total_ibnr': total_ult - total_latest,
        }

    return pd.DataFrame(results).T


# =============================================================================
# Excel Report Generation
# =============================================================================

def write_report(output_path, triangle, acc_periods, dev_periods,
                 individual_factors, selected_factors, tail_factor, tail_method,
                 cl_vw_results, cl_sa_results, cdfs_vw, cdfs_sa,
                 bf_results, cc_results, cc_elr,
                 cal_year_diag, outlier_df, tail_sens,
                 loss_type="incurred", premiums=None):
    """Write formatted Excel report with all exhibits."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, size=11, name='Arial')
    title_font = Font(bold=True, size=14, name='Arial')
    subtitle_font = Font(bold=True, size=12, name='Arial', color='333333')
    num_fmt = '#,##0'
    pct_fmt = '0.0%'
    factor_fmt = '0.0000'
    border_thin = Border(
        bottom=Side(style='thin', color='AAAAAA')
    )
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font_white = Font(bold=True, size=11, name='Arial', color='FFFFFF')
    light_fill = PatternFill('solid', fgColor='D6E4F0')
    warning_fill = PatternFill('solid', fgColor='FFC7CE')
    good_fill = PatternFill('solid', fgColor='C6EFCE')

    def style_header_row(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def auto_width(ws, max_col, min_width=12):
        for c in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = max(min_width, 15)

    # ---- Exhibit 1: Triangle ----
    ws = wb.active
    ws.title = "1-Triangle"
    ws['A1'] = f"Exhibit 1: {loss_type.title()} Loss Development Triangle"
    ws['A1'].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dev_periods) + 1)

    # Headers
    ws.cell(row=3, column=1, value="Accident Period")
    for j, d in enumerate(dev_periods):
        ws.cell(row=3, column=j + 2, value=d)
    style_header_row(ws, 3, len(dev_periods) + 1)

    # Data
    for i, ap in enumerate(acc_periods):
        ws.cell(row=i + 4, column=1, value=ap)
        ws.cell(row=i + 4, column=1).font = Font(bold=True, name='Arial')
        for j in range(len(dev_periods)):
            val = triangle.iloc[i, j]
            if pd.notna(val):
                cell = ws.cell(row=i + 4, column=j + 2, value=val)
                cell.number_format = num_fmt

    auto_width(ws, len(dev_periods) + 1)

    # ---- Exhibit 2: Age-to-Age Factors ----
    ws2 = wb.create_sheet("2-ATA Factors")
    ws2['A1'] = "Exhibit 2: Age-to-Age Development Factors"
    ws2['A1'].font = title_font
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dev_periods))

    factor_cols = dev_periods[:-1]
    ws2.cell(row=3, column=1, value="Accident Period")
    for j, d in enumerate(factor_cols):
        ws2.cell(row=3, column=j + 2, value=f"{d}-{dev_periods[j+1]}")
    style_header_row(ws2, 3, len(factor_cols) + 1)

    for i, ap in enumerate(acc_periods):
        ws2.cell(row=i + 4, column=1, value=ap)
        ws2.cell(row=i + 4, column=1).font = Font(bold=True, name='Arial')
        for j in range(len(factor_cols)):
            val = individual_factors.iloc[i, j]
            if pd.notna(val):
                cell = ws2.cell(row=i + 4, column=j + 2, value=val)
                cell.number_format = factor_fmt

    # Selected factors summary
    summary_start = len(acc_periods) + 5
    for label_idx, (label, method_key) in enumerate([
        ("Volume-Weighted", "volume_weighted"),
        ("Simple Average", "simple_avg"),
        ("Medial Average", "medial_avg"),
    ]):
        row = summary_start + label_idx
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=1).font = Font(bold=True, name='Arial')
        ws2.cell(row=row, column=1).fill = light_fill
        for j, d in enumerate(factor_cols):
            val = selected_factors.get(d, {}).get(method_key, np.nan)
            if pd.notna(val):
                cell = ws2.cell(row=row, column=j + 2, value=val)
                cell.number_format = factor_fmt
                cell.fill = light_fill

    # Selected and tail
    sel_row = summary_start + 4
    ws2.cell(row=sel_row, column=1, value="SELECTED (Vol-Wtd)")
    ws2.cell(row=sel_row, column=1).font = Font(bold=True, name='Arial', color='2F5496')
    for j, d in enumerate(factor_cols):
        val = selected_factors.get(d, {}).get('volume_weighted', np.nan)
        if pd.notna(val):
            cell = ws2.cell(row=sel_row, column=j + 2, value=val)
            cell.number_format = factor_fmt
            cell.font = Font(bold=True, name='Arial', color='2F5496')

    ws2.cell(row=sel_row + 1, column=1, value="Tail Factor")
    ws2.cell(row=sel_row + 1, column=1).font = Font(bold=True, name='Arial')
    ws2.cell(row=sel_row + 1, column=2, value=tail_factor)
    ws2.cell(row=sel_row + 1, column=2).number_format = factor_fmt
    ws2.cell(row=sel_row + 2, column=1, value=f"Tail method: {tail_method}")
    ws2.cell(row=sel_row + 2, column=1).font = Font(italic=True, name='Arial', color='666666')

    auto_width(ws2, len(factor_cols) + 1)

    # ---- Exhibit 3: CDF and Ultimates ----
    ws3 = wb.create_sheet("3-CL Ultimates")
    ws3['A1'] = "Exhibit 3: Chain Ladder Ultimate Losses"
    ws3['A1'].font = title_font
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    cols = ["Accident Period", "Latest Loss", "Dev Age", "CDF",
            "% Reported", "CL Ultimate (VW)", "IBNR (VW)", "CL Ultimate (SA)"]
    for j, c in enumerate(cols):
        ws3.cell(row=3, column=j + 1, value=c)
    style_header_row(ws3, 3, len(cols))

    total_latest = 0
    total_ult_vw = 0
    total_ibnr_vw = 0
    total_ult_sa = 0

    for i in range(len(cl_vw_results)):
        row = i + 4
        vw = cl_vw_results.iloc[i]
        sa = cl_sa_results.iloc[i]
        ws3.cell(row=row, column=1, value=vw['accident_period'])
        ws3.cell(row=row, column=1).font = Font(bold=True, name='Arial')

        for col, val, fmt in [
            (2, vw['latest_loss'], num_fmt),
            (3, vw['latest_dev'], '0'),
            (4, vw['cdf'], factor_fmt),
            (5, vw['pct_reported'], pct_fmt),
            (6, vw['ultimate'], num_fmt),
            (7, vw['ibnr'], num_fmt),
            (8, sa['ultimate'], num_fmt),
        ]:
            if pd.notna(val):
                cell = ws3.cell(row=row, column=col, value=val)
                cell.number_format = fmt
                if col == 7 and val < 0:
                    cell.fill = warning_fill

        if pd.notna(vw['latest_loss']):
            total_latest += vw['latest_loss']
        if pd.notna(vw['ultimate']):
            total_ult_vw += vw['ultimate']
        if pd.notna(vw['ibnr']):
            total_ibnr_vw += vw['ibnr']
        if pd.notna(sa['ultimate']):
            total_ult_sa += sa['ultimate']

    # Totals
    total_row = len(cl_vw_results) + 4
    ws3.cell(row=total_row, column=1, value="TOTAL")
    ws3.cell(row=total_row, column=1).font = Font(bold=True, name='Arial')
    ws3.cell(row=total_row, column=1).fill = light_fill
    for col, val in [(2, total_latest), (6, total_ult_vw), (7, total_ibnr_vw), (8, total_ult_sa)]:
        cell = ws3.cell(row=total_row, column=col, value=val)
        cell.number_format = num_fmt
        cell.font = Font(bold=True, name='Arial')
        cell.fill = light_fill

    auto_width(ws3, len(cols))

    # ---- Exhibit 4: BF and Cape Cod (if available) ----
    if bf_results is not None or cc_results is not None:
        ws4 = wb.create_sheet("4-BF Cape Cod")
        ws4['A1'] = "Exhibit 4: Bornhuetter-Ferguson & Cape Cod Ultimates"
        ws4['A1'].font = title_font
        ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        bf_cols = ["Accident Period", "Latest Loss", "Earned Premium",
                   "BF ELR", "% Unreported", "BF IBNR", "BF Ultimate",
                   "CC IBNR", "CC Ultimate"]
        for j, c in enumerate(bf_cols):
            ws4.cell(row=3, column=j + 1, value=c)
        style_header_row(ws4, 3, len(bf_cols))

        n_rows = len(bf_results) if bf_results is not None else len(cc_results)
        for i in range(n_rows):
            row = i + 4
            if bf_results is not None:
                bf = bf_results.iloc[i]
                ws4.cell(row=row, column=1, value=bf['accident_period'])
                ws4.cell(row=row, column=1).font = Font(bold=True, name='Arial')
                for col, val, fmt in [
                    (2, bf['latest_loss'], num_fmt),
                    (3, bf['premium'], num_fmt),
                    (4, bf['elr'], pct_fmt),
                    (5, bf['pct_unreported'], pct_fmt),
                    (6, bf['bf_ibnr'], num_fmt),
                    (7, bf['ultimate'], num_fmt),
                ]:
                    if pd.notna(val):
                        cell = ws4.cell(row=row, column=col, value=val)
                        cell.number_format = fmt

            if cc_results is not None:
                cc = cc_results.iloc[i]
                for col, val, fmt in [
                    (8, cc['cc_ibnr'], num_fmt),
                    (9, cc['ultimate'], num_fmt),
                ]:
                    if pd.notna(val):
                        cell = ws4.cell(row=row, column=col, value=val)
                        cell.number_format = fmt

        if cc_elr is not None:
            note_row = n_rows + 5
            ws4.cell(row=note_row, column=1, value=f"Cape Cod derived ELR: {cc_elr:.1%}")
            ws4.cell(row=note_row, column=1).font = Font(italic=True, name='Arial', color='666666')

        auto_width(ws4, len(bf_cols))

    # ---- Exhibit 5: Diagnostics ----
    ws5 = wb.create_sheet("5-Diagnostics")
    ws5['A1'] = "Exhibit 5: Diagnostic Tests"
    ws5['A1'].font = title_font
    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # Calendar year test
    ws5['A3'] = "Calendar Year Development Test"
    ws5['A3'].font = subtitle_font
    diag_cols = ["Diagonal", "Sum", "Prior Sum", "Change"]
    for j, c in enumerate(diag_cols):
        ws5.cell(row=4, column=j + 1, value=c)
    style_header_row(ws5, 4, len(diag_cols))

    if not cal_year_diag.empty:
        for i in range(len(cal_year_diag)):
            row = i + 5
            d = cal_year_diag.iloc[i]
            ws5.cell(row=row, column=1, value=d['diagonal'])
            ws5.cell(row=row, column=2, value=d['sum'])
            ws5.cell(row=row, column=2).number_format = num_fmt
            ws5.cell(row=row, column=3, value=d['prior_sum'])
            ws5.cell(row=row, column=3).number_format = num_fmt
            if pd.notna(d['change']):
                cell = ws5.cell(row=row, column=4, value=d['change'])
                cell.number_format = pct_fmt
                if abs(d['change']) > 0.15:
                    cell.fill = warning_fill

    # Outlier factors
    outlier_start = len(cal_year_diag) + 7 if not cal_year_diag.empty else 7
    ws5.cell(row=outlier_start, column=1, value="Outlier Age-to-Age Factors (>2σ)")
    ws5.cell(row=outlier_start, column=1).font = subtitle_font

    if not outlier_df.empty:
        out_cols = ["Dev Period", "Row", "Factor", "Mean", "Std Devs"]
        for j, c in enumerate(out_cols):
            ws5.cell(row=outlier_start + 1, column=j + 1, value=c)
        style_header_row(ws5, outlier_start + 1, len(out_cols))

        for i in range(len(outlier_df)):
            row = outlier_start + 2 + i
            o = outlier_df.iloc[i]
            ws5.cell(row=row, column=1, value=o['dev_period'])
            ws5.cell(row=row, column=2, value=o['row_index'])
            ws5.cell(row=row, column=3, value=o['factor'])
            ws5.cell(row=row, column=3).number_format = factor_fmt
            ws5.cell(row=row, column=4, value=o['mean'])
            ws5.cell(row=row, column=4).number_format = factor_fmt
            ws5.cell(row=row, column=5, value=o['std_devs'])
            ws5.cell(row=row, column=5).number_format = '0.00'
    else:
        ws5.cell(row=outlier_start + 1, column=1, value="No outlier factors detected.")
        ws5.cell(row=outlier_start + 1, column=1).font = Font(italic=True, name='Arial', color='006600')

    # Tail sensitivity
    tail_start = outlier_start + (len(outlier_df) if not outlier_df.empty else 0) + 4
    ws5.cell(row=tail_start, column=1, value="Tail Factor Sensitivity")
    ws5.cell(row=tail_start, column=1).font = subtitle_font

    sens_cols = ["Scenario", "Tail Factor", "Total Ultimate", "Total IBNR"]
    for j, c in enumerate(sens_cols):
        ws5.cell(row=tail_start + 1, column=j + 1, value=c)
    style_header_row(ws5, tail_start + 1, len(sens_cols))

    for i, (label, row_data) in enumerate(tail_sens.iterrows()):
        row = tail_start + 2 + i
        ws5.cell(row=row, column=1, value=label)
        ws5.cell(row=row, column=2, value=row_data['tail_factor'])
        ws5.cell(row=row, column=2).number_format = factor_fmt
        ws5.cell(row=row, column=3, value=row_data['total_ultimate'])
        ws5.cell(row=row, column=3).number_format = num_fmt
        ws5.cell(row=row, column=4, value=row_data['total_ibnr'])
        ws5.cell(row=row, column=4).number_format = num_fmt
        if label == 'Selected':
            for c in range(1, 5):
                ws5.cell(row=row, column=c).font = Font(bold=True, name='Arial')

    auto_width(ws5, 5)

    # ---- Exhibit 6: Summary Comparison ----
    ws6 = wb.create_sheet("6-Summary")
    ws6['A1'] = "Exhibit 6: Method Comparison Summary"
    ws6['A1'].font = title_font
    ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    sum_cols = ["Accident Period", "Latest Loss", "CL (VW)", "CL (SA)"]
    if bf_results is not None:
        sum_cols.append("BF")
    if cc_results is not None:
        sum_cols.append("Cape Cod")
    sum_cols.extend(["Range", "Range %"])

    for j, c in enumerate(sum_cols):
        ws6.cell(row=3, column=j + 1, value=c)
    style_header_row(ws6, 3, len(sum_cols))

    for i in range(len(cl_vw_results)):
        row = i + 4
        vw_ult = cl_vw_results.iloc[i]['ultimate']
        sa_ult = cl_sa_results.iloc[i]['ultimate']
        latest = cl_vw_results.iloc[i]['latest_loss']

        ws6.cell(row=row, column=1, value=cl_vw_results.iloc[i]['accident_period'])
        ws6.cell(row=row, column=1).font = Font(bold=True, name='Arial')
        ws6.cell(row=row, column=2, value=latest)
        ws6.cell(row=row, column=2).number_format = num_fmt
        ws6.cell(row=row, column=3, value=vw_ult)
        ws6.cell(row=row, column=3).number_format = num_fmt
        ws6.cell(row=row, column=4, value=sa_ult)
        ws6.cell(row=row, column=4).number_format = num_fmt

        ults = [v for v in [vw_ult, sa_ult] if pd.notna(v)]
        col_offset = 5

        if bf_results is not None:
            bf_ult = bf_results.iloc[i]['ultimate']
            ws6.cell(row=row, column=col_offset, value=bf_ult)
            ws6.cell(row=row, column=col_offset).number_format = num_fmt
            if pd.notna(bf_ult):
                ults.append(bf_ult)
            col_offset += 1

        if cc_results is not None:
            cc_ult = cc_results.iloc[i]['ultimate']
            ws6.cell(row=row, column=col_offset, value=cc_ult)
            ws6.cell(row=row, column=col_offset).number_format = num_fmt
            if pd.notna(cc_ult):
                ults.append(cc_ult)
            col_offset += 1

        if len(ults) >= 2:
            rng = max(ults) - min(ults)
            rng_pct = rng / min(ults) if min(ults) != 0 else np.nan
            ws6.cell(row=row, column=col_offset, value=rng)
            ws6.cell(row=row, column=col_offset).number_format = num_fmt
            ws6.cell(row=row, column=col_offset + 1, value=rng_pct)
            ws6.cell(row=row, column=col_offset + 1).number_format = pct_fmt
            if pd.notna(rng_pct) and rng_pct > 0.10:
                ws6.cell(row=row, column=col_offset + 1).fill = warning_fill

    auto_width(ws6, len(sum_cols))

    wb.save(output_path)
    return output_path


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Reserve Adequacy Quick Check")
    parser.add_argument("--input", required=True, help="Path to triangle file")
    parser.add_argument("--sheet", default=None, help="Sheet name for Excel")
    parser.add_argument("--type", default="incurred", choices=["paid", "incurred", "both"])
    parser.add_argument("--dev-periods", default="years", choices=["months", "years"])
    parser.add_argument("--premium", default=None, help="Path to premium file or JSON dict")
    parser.add_argument("--elr", type=float, default=None, help="A priori expected loss ratio for BF")
    parser.add_argument("--output", default="reserve_report.xlsx", help="Output Excel path")
    args = parser.parse_args()

    # Parse triangle
    print("Parsing triangle...")
    acc_periods, dev_periods, triangle, metadata = read_triangle(args.input, args.sheet)
    print(json.dumps(metadata, indent=2))

    # Parse premium if provided
    premiums = None
    if args.premium:
        try:
            premiums = json.loads(args.premium)
        except (json.JSONDecodeError, TypeError):
            if os.path.isfile(args.premium):
                premiums = read_premium(args.premium)

    # Compute age-to-age factors
    print("\nComputing age-to-age factors...")
    individual_factors, selected_factors = compute_ata_factors(triangle)

    # Tail factor
    print("Estimating tail factor...")
    tail_factor, tail_method = compute_tail_factor(selected_factors, dev_periods)
    print(f"  Tail: {tail_factor:.4f} ({tail_method})")

    # CDFs
    cdfs_vw = compute_cdfs(selected_factors, dev_periods, tail_factor, "volume_weighted")
    cdfs_sa = compute_cdfs(selected_factors, dev_periods, tail_factor, "simple_avg")

    # Chain Ladder ultimates
    print("Running Chain Ladder (Volume-Weighted)...")
    cl_vw = chain_ladder_ultimates(triangle, acc_periods, cdfs_vw, dev_periods)
    print("Running Chain Ladder (Simple Average)...")
    cl_sa = chain_ladder_ultimates(triangle, acc_periods, cdfs_sa, dev_periods)

    # BF and Cape Cod
    bf_results = None
    cc_results = None
    cc_elr = None
    if premiums:
        # Determine ELR for BF
        if args.elr:
            elr = args.elr
        else:
            # Derive from most mature chain ladder results
            mature_ults = []
            mature_prems = []
            for _, row in cl_vw.iterrows():
                prem = premiums.get(str(row['accident_period']), np.nan)
                if pd.notna(row['ultimate']) and pd.notna(prem) and prem > 0:
                    if pd.notna(row['pct_reported']) and row['pct_reported'] > 0.8:
                        mature_ults.append(row['ultimate'])
                        mature_prems.append(prem)
            elr = sum(mature_ults) / sum(mature_prems) if mature_prems else 0.65
            print(f"  Derived BF ELR from mature years: {elr:.1%}")

        print("Running Bornhuetter-Ferguson...")
        bf_results = bornhuetter_ferguson(triangle, acc_periods, dev_periods, cdfs_vw, premiums, elr)

        print("Running Cape Cod...")
        cc_results, cc_elr = cape_cod(triangle, acc_periods, dev_periods, cdfs_vw, premiums)
        print(f"  Cape Cod ELR: {cc_elr:.1%}")

    # Diagnostics
    print("Running diagnostics...")
    cal_diag = calendar_year_test(triangle, acc_periods, dev_periods)
    outliers = outlier_factors(individual_factors, selected_factors, dev_periods)
    tail_sens = tail_sensitivity(cl_vw, cdfs_vw, tail_factor, dev_periods)

    # Write report
    print(f"\nWriting report to {args.output}...")
    write_report(
        args.output, triangle, acc_periods, dev_periods,
        individual_factors, selected_factors, tail_factor, tail_method,
        cl_vw, cl_sa, cdfs_vw, cdfs_sa,
        bf_results, cc_results, cc_elr,
        cal_diag, outliers, tail_sens,
        loss_type=args.type, premiums=premiums
    )

    # Print summary
    print("\n" + "=" * 60)
    print("RESERVE ADEQUACY QUICK CHECK — SUMMARY")
    print("=" * 60)
    total_latest = cl_vw['latest_loss'].sum()
    total_cl_vw = cl_vw['ultimate'].sum()
    total_cl_sa = cl_sa['ultimate'].sum()
    total_ibnr_vw = cl_vw['ibnr'].sum()

    print(f"Total Latest {args.type.title()} Losses: {total_latest:,.0f}")
    print(f"CL Ultimate (Vol-Wtd):  {total_cl_vw:,.0f}  |  IBNR: {total_ibnr_vw:,.0f}")
    print(f"CL Ultimate (Simple):   {total_cl_sa:,.0f}")
    if bf_results is not None:
        total_bf = bf_results['ultimate'].sum()
        print(f"BF Ultimate:            {total_bf:,.0f}")
    if cc_results is not None:
        total_cc = cc_results['ultimate'].sum()
        print(f"Cape Cod Ultimate:      {total_cc:,.0f}")

    print(f"\nTail Factor: {tail_factor:.4f} ({tail_method})")
    if not outliers.empty:
        print(f"Outlier factors detected: {len(outliers)}")
    else:
        print("No outlier age-to-age factors detected.")
    print(f"\nReport saved to: {args.output}")


if __name__ == "__main__":
    main()
